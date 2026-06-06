#!/usr/bin/env python3
"""Normalize mixed contact sources into one standard intermediate CSV."""

from __future__ import annotations

import argparse
import csv
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

from source_format_utils import detect_header_row, guess_source_format, read_csv_rows, safe_stem


PROJECT_ROOT = Path(__file__).resolve().parents[1]
STANDARD_FIELDS = [
    "source_file",
    "source_sheet",
    "source_row_number",
    "source_format",
    "raw_name",
    "raw_phone",
    "raw_email",
    "raw_notes",
    "raw_payload_json",
    "raw_phones_json",
    "raw_emails_json",
    "aliases_json",
    "organization",
    "job_title",
    "source_category",
    "campaign_name",
    "lead_status",
    "requirement_json",
    "inventory_hint_json",
    "building_name_hint",
    "building_code_hint",
    "wing_hint",
    "unit_number_hint",
    "relationship_hint",
    "contact_person_1",
    "contact_person_2",
    "source_label",
    "website",
    "google_maps_link",
]

PHONE_COLUMNS = [
    "Phone",
    "Phone Number",
    "Phone 2",
    "Mobile",
    "Mobile Phone",
    "Home Phone",
    "Business Phone",
    "Home Fax",
    "Business Fax",
    "Pager",
    "Telephone 1",
    "Telephone 2",
    "Contact Number",
    "contact number",
    "Mobile Number",
    "phone_number",
]
EMAIL_COLUMNS = ["Email", "E-mail Address", "E-mail 2 Address", "E-mail 3 Address", "email"]


def row_get(row: Dict[str, object], *keys: str) -> str:
    normalized = {str(key).strip().lower(): value for key, value in row.items()}
    for key in keys:
        value = normalized.get(key.strip().lower())
        if value is not None:
            return str(value).strip()
    return ""


def json_list(values: Iterable[str]) -> str:
    items = []
    seen = set()
    for value in values:
        cleaned = str(value or "").strip()
        if cleaned and cleaned.lower() not in {"n/a", "na", "none", "null", "-"} and cleaned not in seen:
            items.append(cleaned)
            seen.add(cleaned)
    return json.dumps(items, ensure_ascii=True)


def json_payload(row: Dict[str, object]) -> str:
    return json.dumps({str(key): "" if value is None else str(value) for key, value in row.items()}, ensure_ascii=True, sort_keys=True)


def json_object(values: Dict[str, object]) -> str:
    cleaned = {}
    for key, value in values.items():
        if value is None or value == "":
            continue
        if isinstance(value, list) and not value:
            continue
        cleaned[key] = value
    return json.dumps(cleaned, ensure_ascii=True, sort_keys=True)


def collect_columns(row: Dict[str, object], columns: Sequence[str]) -> List[str]:
    return [row_get(row, column) for column in columns if row_get(row, column)]


def make_standard_row(
    source_file: Path,
    source_sheet: str,
    source_row_number: int,
    source_format: str,
    row: Dict[str, object],
    **values: str,
) -> Dict[str, str]:
    phones = collect_columns(row, PHONE_COLUMNS)
    emails = collect_columns(row, EMAIL_COLUMNS)
    output = {field: "" for field in STANDARD_FIELDS}
    output.update(
        {
            "source_file": str(source_file),
            "source_sheet": source_sheet,
            "source_row_number": str(source_row_number),
            "source_format": source_format,
            "raw_payload_json": json_payload(row),
            "raw_phones_json": json_list(phones),
            "raw_emails_json": json_list(emails),
        }
    )
    for key, value in values.items():
        output[key] = value or ""
    if not output["raw_phone"] and phones:
        output["raw_phone"] = phones[0]
    if not output["raw_email"] and emails:
        output["raw_email"] = emails[0]
    return output


def display_name(row: Dict[str, object]) -> str:
    return row_get(row, "Display Name") or " ".join(part for part in [row_get(row, "First Name"), row_get(row, "Last Name")] if part).strip()


def normalize_simple(source_file: Path, sheet: str, row_number: int, row: Dict[str, object], source_format: str) -> List[Dict[str, str]]:
    return [
        make_standard_row(
            source_file,
            sheet,
            row_number,
            source_format,
            row,
            raw_name=row_get(row, "Name", "Display Name", "Title"),
            raw_phone=row_get(row, "Phone Number", "Phone", "Mobile"),
            raw_email=row_get(row, "Email"),
            raw_notes=row_get(row, "Comment", "Notes"),
            source_label=row_get(row, "Source"),
        )
    ]


def normalize_google_contacts(source_file: Path, sheet: str, row_number: int, row: Dict[str, object], source_format: str) -> List[Dict[str, str]]:
    aliases = [row_get(row, "First Name"), row_get(row, "Last Name"), row_get(row, "Display Name")]
    return [
        make_standard_row(
            source_file,
            sheet,
            row_number,
            source_format,
            row,
            raw_name=display_name(row),
            raw_phone=row_get(row, "Mobile Phone", "Home Phone", "Business Phone"),
            raw_email=row_get(row, "E-mail Address", "E-mail 2 Address", "E-mail 3 Address"),
            raw_notes=row_get(row, "Notes"),
            aliases_json=json_list(aliases),
            organization=row_get(row, "Organization"),
            job_title=row_get(row, "Job Title"),
            source_category=row_get(row, "Categories"),
        )
    ]


def normalize_whatsapp(source_file: Path, sheet: str, row_number: int, row: Dict[str, object], source_format: str) -> List[Dict[str, str]]:
    push_name = row_get(row, "Push Name")
    return [
        make_standard_row(
            source_file,
            sheet,
            row_number,
            source_format,
            row,
            raw_name=row_get(row, "Name") or push_name,
            raw_phone=row_get(row, "number"),
            raw_notes=f"push_name_present={bool(push_name)}",
            aliases_json=json_list([push_name]),
            source_label="whatsapp",
        )
    ]


def normalize_messy(source_file: Path, sheet: str, row_number: int, row: Dict[str, object], source_format: str) -> List[Dict[str, str]]:
    rows = normalize_simple(source_file, sheet, row_number, row, source_format)
    later_name = row_get(row, "N", "Name 2", "Contact Person", "Contact Person 1")
    later_phone = row_get(row, "Telephone 1", "Phone 2", "Mobile 1")
    flat = row_get(row, "Flat No.", "Flat", "Unit")
    if later_name or later_phone or flat:
        rows.append(
            make_standard_row(
                source_file,
                sheet,
                row_number,
                source_format,
                row,
                raw_name=later_name,
                raw_phone=later_phone,
                wing_hint=row_get(row, "Wing"),
                unit_number_hint=flat,
                relationship_hint=row_get(row, "Relationship", "Role"),
                source_label=row_get(row, "Source"),
            )
        )
    return rows


def normalize_structured(source_file: Path, sheet: str, row_number: int, row: Dict[str, object], source_format: str) -> List[Dict[str, str]]:
    output: List[Dict[str, str]] = []
    wing = row_get(row, "Wing", "Tower", "TOWER")
    flat = row_get(row, "Flat No.", "Flat No", "FLAT NO.", "FLAT NO", "Flat", "Unit", "FLAT DETAILS", "Premises Details")
    email = row_get(row, "Email", "E-mail", "E-MAIL", "EMAIL ID", "EMAID ID", "PRIMARY EMAIL ID")
    source = row_get(row, "Source")
    person_1 = row_get(row, "Contact Person 1", "Name", "NAME", "CLIENT NAME", "Member Name", "Name of Lessee", "NAME OF THE PARTY")
    person_2 = row_get(row, "Contact Person 2")
    phone_1 = row_get(row, "Telephone 1", "Phone 1", "Mobile 1", "Contact Number", "CONTACT DETAILS", "CONTACT NO.", "CONTACT NO", "PHONE NO", "PRIMARY CONTACT NO", "Mobile No.", "CUSTOMERS NUMBER", "Number")
    phone_2 = row_get(row, "Telephone 2", "Phone 2", "Mobile 2", "SECONDARY CONTACT NO", "OTHER CONTACT NO/S.", "BROKERS", "BROKER")
    building = sheet if source_format in {"owner_tenant_all_projects_workbook", "multi_sheet_project_contacts_workbook"} else ""
    member = {
        "member_type": row_get(row, "Member Type"),
        "occupied_by": row_get(row, "Occupied By"),
        "address_present": bool(row_get(row, "Address", "Address 1", "Address 2")),
        "typology": row_get(row, "CONFIG", "TYPE", "Typology"),
        "carpet_area": row_get(row, "carpet area", "CARPET AREA"),
        "broker_present": bool(row_get(row, "BROKER", "Broker")),
    }

    for person, phone in [(person_1, phone_1), (person_2, phone_2)]:
        if not (person or phone):
            continue
        output.append(
            make_standard_row(
                source_file,
                sheet,
                row_number,
                source_format,
                row,
                raw_name=person,
                raw_phone=phone,
                raw_email=email,
                building_name_hint=building,
                wing_hint=wing,
                unit_number_hint=flat,
                relationship_hint=row_get(row, "Type", "Relationship", "Member Type") or ("broker" if "broker" in source_file.name.lower() else "owner"),
                inventory_hint_json=json_object(member),
                contact_person_1=person_1,
                contact_person_2=person_2,
                source_label=source,
            )
        )
    return output


def normalize_google_maps(source_file: Path, sheet: str, row_number: int, row: Dict[str, object], source_format: str) -> List[Dict[str, str]]:
    notes = "; ".join(
        item
        for item in [
            f"industry={row_get(row, 'Industry')}" if row_get(row, "Industry") else "",
            f"rating_present={bool(row_get(row, 'Rating'))}",
            f"reviews_present={bool(row_get(row, 'Reviews'))}",
            f"address_present={bool(row_get(row, 'Address'))}",
        ]
        if item
    )
    return [
        make_standard_row(
            source_file,
            sheet,
            row_number,
            source_format,
            row,
            raw_name=row_get(row, "Title"),
            raw_phone=row_get(row, "Phone"),
            raw_notes=notes,
            relationship_hint="business_lead",
            source_label=row_get(row, "Industry"),
            website=row_get(row, "Website"),
            google_maps_link=row_get(row, "Google Maps Link"),
        )
    ]


def normalize_meta_lead(source_file: Path, sheet: str, row_number: int, row: Dict[str, object], source_format: str) -> List[Dict[str, str]]:
    return [
        make_standard_row(
            source_file,
            sheet,
            row_number,
            source_format,
            row,
            raw_name=row_get(row, "full_name", "name"),
            raw_phone=row_get(row, "phone_number", "phone"),
            raw_email=row_get(row, "email"),
            relationship_hint="lead",
            campaign_name=row_get(row, "campaign_name", "campaign_id"),
            lead_status=row_get(row, "lead_status"),
            source_label=row_get(row, "platform"),
        )
    ]


def normalize_portal_lead(source_file: Path, sheet: str, row_number: int, row: Dict[str, object], source_format: str) -> List[Dict[str, str]]:
    requirement = {
        "property_type": row_get(row, "PropertyType", "Property Type"),
        "purpose": row_get(row, "Purpose"),
        "locality": row_get(row, "Locality"),
        "city": row_get(row, "City"),
        "budget": row_get(row, "Budget"),
        "visit": row_get(row, "Visit"),
    }
    return [
        make_standard_row(
            source_file,
            sheet,
            row_number,
            source_format,
            row,
            raw_name=row_get(row, "Name"),
            raw_phone=row_get(row, "Mobile", "Phone"),
            raw_email=row_get(row, "Email"),
            raw_notes="portal lead requirement present",
            relationship_hint="lead",
            requirement_json=json_object(requirement),
            source_label=row_get(row, "Source", "Portal"),
        )
    ]


def normalize_inventory(source_file: Path, sheet: str, row_number: int, row: Dict[str, object], source_format: str) -> List[Dict[str, str]]:
    phone = row_get(row, "contact number", "Contact Number", "Phone", "Mobile")
    name = row_get(row, "Name", "Owner", "Contact Person")
    inventory = {
        "building_name": row_get(row, "Building Name") or sheet,
        "wing": row_get(row, "Wing"),
        "unit_number": row_get(row, "Flat Number", "Flat No.", "Unit"),
        "typology": row_get(row, "Typology", "Bedrooms"),
        "sqft": row_get(row, "SQFT", "Carpet Area", "Salable Area"),
        "rent": row_get(row, "Rent", "Rent Price"),
        "sale_price": row_get(row, "Sell", "Sale Price", "CP"),
        "for": row_get(row, "For", "Purpose"),
    }
    if not (phone or name):
        return []
    return [
        make_standard_row(
            source_file,
            sheet,
            row_number,
            source_format,
            row,
            raw_name=name or row_get(row, "Building Name") or sheet,
            raw_phone=phone,
            building_name_hint=str(inventory.get("building_name") or ""),
            wing_hint=str(inventory.get("wing") or ""),
            unit_number_hint=str(inventory.get("unit_number") or ""),
            relationship_hint="inventory_contact",
            inventory_hint_json=json_object(inventory),
            source_label="inventory",
        )
    ]


def normalize_unknown(source_file: Path, sheet: str, row_number: int, row: Dict[str, object], source_format: str) -> List[Dict[str, str]]:
    return [
        make_standard_row(
            source_file,
            sheet,
            row_number,
            source_format,
            row,
            raw_name=row_get(row, "Name", "Title", "Contact Person 1", "Display Name"),
            raw_phone=row_get(row, "Phone", "Phone Number", "Mobile", "Telephone 1"),
            raw_email=row_get(row, "Email", "E-mail Address"),
            raw_notes=row_get(row, "Notes", "Source", "Comment"),
            source_label=row_get(row, "Source"),
        )
    ]


def normalize_row(source_file: Path, sheet: str, row_number: int, row: Dict[str, object], source_format: str) -> List[Dict[str, str]]:
    if source_format == "google_contacts_csv":
        return normalize_google_contacts(source_file, sheet, row_number, row, source_format)
    if source_format == "whatsapp_export_csv":
        return normalize_whatsapp(source_file, sheet, row_number, row, source_format)
    if source_format == "simple_phonebook_csv":
        return normalize_simple(source_file, sheet, row_number, row, source_format)
    if source_format == "messy_phonebook_property_csv":
        return normalize_messy(source_file, sheet, row_number, row, source_format)
    if source_format in {"structured_owner_sheet", "structured_owner_workbook", "owner_tenant_all_projects_workbook", "multi_sheet_project_contacts_workbook", "broker_list_workbook", "building_owner_tenant_workbook", "unit_resident_workbook", "project_customer_workbook", "imperial_unit_inventory_workbook", "society_member_details_workbook"}:
        return normalize_structured(source_file, sheet, row_number, row, source_format)
    if source_format == "google_maps_business_csv":
        return normalize_google_maps(source_file, sheet, row_number, row, source_format)
    if source_format == "meta_facebook_leads_utf16_tsv":
        return normalize_meta_lead(source_file, sheet, row_number, row, source_format)
    if source_format == "portal_property_leads_csv":
        return normalize_portal_lead(source_file, sheet, row_number, row, source_format)
    if source_format in {"property_inventory_csv", "property_inventory_workbook"}:
        return normalize_inventory(source_file, sheet, row_number, row, source_format)
    return normalize_unknown(source_file, sheet, row_number, row, source_format)


def parse_vcards(path: Path) -> List[Tuple[int, Dict[str, object]]]:
    cards: List[Tuple[int, Dict[str, object]]] = []
    current: List[str] = []
    card_number = 0
    with path.open(encoding="utf-8", errors="replace") as handle:
        for line in handle:
            stripped = line.rstrip("\n")
            if stripped.upper().startswith("BEGIN:VCARD"):
                current = [stripped]
                card_number += 1
            elif current:
                current.append(stripped)
                if stripped.upper().startswith("END:VCARD"):
                    row: Dict[str, object] = {"_raw_vcard": "\n".join(current)}
                    phones: List[str] = []
                    emails: List[str] = []
                    for item in current:
                        upper = item.upper()
                        if upper.startswith("FN"):
                            row["FN"] = item.split(":", 1)[-1]
                        elif upper.startswith("N"):
                            row["N"] = item.split(":", 1)[-1].replace(";", " ").strip()
                        elif upper.startswith("TEL"):
                            phones.append(item.split(":", 1)[-1])
                        elif upper.startswith("EMAIL"):
                            emails.append(item.split(":", 1)[-1])
                        elif upper.startswith("ORG"):
                            row["ORG"] = item.split(":", 1)[-1]
                        elif upper.startswith("NOTE"):
                            row["NOTE"] = item.split(":", 1)[-1]
                    row["phones"] = phones
                    row["emails"] = emails
                    cards.append((card_number, row))
                    current = []
    return cards


def normalize_vcf(path: Path) -> List[Dict[str, str]]:
    output = []
    for card_number, row in parse_vcards(path):
        phones = [str(item) for item in row.get("phones", []) if str(item)]
        emails = [str(item) for item in row.get("emails", []) if str(item)]
        base = make_standard_row(
            path,
            "",
            card_number,
            "vcf_contacts_file",
            row,
            raw_name=str(row.get("FN") or row.get("N") or ""),
            raw_phone=phones[0] if phones else "",
            raw_email=emails[0] if emails else "",
            raw_notes=str(row.get("NOTE") or ""),
            raw_phones_json=json_list(phones),
            raw_emails_json=json_list(emails),
            organization=str(row.get("ORG") or ""),
        )
        output.append(base)
    return output


def normalize_csv(path: Path) -> List[Dict[str, str]]:
    rows, encoding, delimiter = read_csv_rows(path)
    columns = list(rows[0].keys()) if rows else []
    source_format = guess_source_format(columns, file_name=path.name, delimiter=delimiter, encoding=encoding)
    output: List[Dict[str, str]] = []
    for index, row in enumerate(rows, start=2):
        output.extend(normalize_row(path, "", index, row, source_format))
    return output


def normalize_xlsx(path: Path) -> List[Dict[str, str]]:
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise SystemExit("XLSX normalization needs openpyxl. Export workbook sheets to CSV or install openpyxl.")

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    output: List[Dict[str, str]] = []
    for sheet in workbook.worksheets:
        all_rows = list(sheet.iter_rows(values_only=True))
        if not all_rows:
            continue
        header_index, columns = detect_header_row(all_rows)
        if not columns:
            continue
        sheet_format = guess_source_format(columns, file_name=path.name, sheet_names=sheet_names)
        if sheet_format == "structured_owner_sheet":
            sheet_format = "structured_owner_workbook"
        for row_number, values in enumerate(all_rows[header_index + 1 :], start=header_index + 2):
            row = {columns[index]: values[index] if index < len(values) else "" for index in range(len(columns)) if columns[index]}
            if not any(str(value or "").strip() for value in row.values()):
                continue
            output.extend(normalize_row(path, sheet.title, row_number, row, sheet_format))
    workbook.close()
    return output


def normalize_text_lines(path: Path, source_format: str) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    line_pattern = re.compile(r"(.+?)\\s+(\\+?\\d[\\d\\s().-]{7,}\\d)")
    with path.open(encoding="utf-8", errors="replace") as handle:
        for row_number, line in enumerate(handle, start=1):
            match = line_pattern.search(line.strip())
            if not match:
                continue
            row = {"line": line.strip(), "Name": match.group(1), "Phone": match.group(2)}
            rows.extend(normalize_simple(path, "", row_number, row, source_format))
    return rows


def normalize_file(path: Path) -> List[Dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv" or path.name.endswith(".csv.example"):
        return normalize_csv(path)
    if suffix == ".xlsx":
        return normalize_xlsx(path)
    if suffix == ".vcf" or path.name.endswith(".vcf.example"):
        return normalize_vcf(path)
    if suffix in {".txt", ".text"}:
        return normalize_text_lines(path, "txt_contacts")
    print("Unsupported or profile-only file type. No normalized rows produced.")
    return []


def write_output(rows: List[Dict[str, str]], source_file: Path, output_dir: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"normalized_contacts_{safe_stem(source_file)}_{timestamp}.csv"
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=STANDARD_FIELDS)
        writer.writeheader()
        writer.writerows(rows)
    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Normalize contact source files into a standard intermediate CSV.")
    parser.add_argument("source_file", help="CSV, XLSX, VCF, or simple TXT source file.")
    parser.add_argument("--output-dir", default=str(PROJECT_ROOT / "exports" / "contacts"), help="Output directory. Defaults to exports/contacts.")
    args = parser.parse_args()

    path = Path(args.source_file)
    if not path.exists():
        print("Source file was not found.")
        return 1

    rows = normalize_file(path)
    output_path = write_output(rows, path, Path(args.output_dir))
    print(f"Normalized rows written: {len(rows)}")
    print(f"Normalized output: {output_path}")
    if len(rows) == 0:
        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".csv"}:
            print("No contact rows normalized. If this is an inventory-only file, use future inventory import.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
