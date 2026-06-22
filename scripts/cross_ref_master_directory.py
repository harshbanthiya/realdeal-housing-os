#!/usr/bin/env python3
"""Phase 6.31 — cross-reference the Kalpataru Master Contact Directory against our data.

The master directory (operator xlsx, scraped from community WhatsApp-group screenshots) is 196
phone numbers + region, with NO names / flats / emails. This:
  1. normalises every master number (Indian 10-digit, else +E.164 intl),
  2. matches each against (a) our flat-tagged rental call list and (b) the DB contact_methods,
  3. labels each: known-with-flat / known-contact / NEW (unknown to us),
  4. emits a cross-ref CSV + a TrueCaller verification WORKLIST (the unknowns, with blank
     name/email columns to fill in the Truecaller app), and an ingest-back template.

No Truecaller API is called (none is available to consumers); the worklist is for manual app
lookup, then `--ingest <filled.csv>` writes the verified names/emails back. Read-only otherwise.
"""

from __future__ import annotations
from _db import psql

import argparse
import csv
import re
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
ENV_FILE = PROJECT_ROOT / "docker" / ".env"
OUT_DIR = PROJECT_ROOT / "exports" / "master_directory"
MASTER_XLSX = Path.home() / "Downloads" / "Kalpataru_Master_Contact_Directory.xlsx"
CALL_LIST = PROJECT_ROOT / "exports" / "rental_all_wings" / "rental_call_list_all_wings.csv"

def env(key: str) -> str:
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
            if line.startswith(f"{key}="):
                return line.split("=", 1)[1]
    return ""
def norm(raw: str) -> str | None:
    s = str(raw or "").strip()
    d = re.sub(r"\D", "", s)
    if not d:
        return None
    if s.startswith("+"):                          # explicit country code -> trust it
        if d.startswith("91") and len(d) == 12:
            return d[2:]                            # +91 -> Indian 10-digit
        return "+" + d                              # any other +CC stays international
    if d.startswith("91") and len(d) == 12:
        d = d[2:]
    if len(d) == 10 and d[0] in "6789":
        return d                                    # bare Indian mobile
    if len(d) >= 11:
        return "+" + d
    return None

def read_master() -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(MASTER_XLSX, read_only=True, data_only=True)
    ws = wb["Master Directory"] if "Master Directory" in wb.sheetnames else wb.worksheets[0]
    out = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        raw = str(r[1]) if len(r) > 1 and r[1] else ""
        n = norm(raw)
        if n:
            out.append({"serial": str(r[0]) if r[0] else "", "raw": raw, "phone": n,
                        "region": str(r[3]) if len(r) > 3 and r[3] else "",
                        "source": str(r[2]) if len(r) > 2 and r[2] else ""})
    return out

def load_known() -> tuple[dict, dict]:
    flat_map: dict[str, dict] = {}
    if CALL_LIST.exists():
        for r in csv.DictReader(CALL_LIST.open()):
            n = norm(r["phone"])
            if n and n not in flat_map:
                flat_map[n] = {"name": r["name_on_file"], "flat": r["flat"], "wing": r["wing"]}
    contact_map: dict[str, str] = {}
    rows = psql("select coalesce(m.normalized_value,m.raw_value), c.full_name from contact_methods m "
                "join contacts c on c.id=m.contact_id where m.method_type in ('mobile','phone');")
    for ln in rows.splitlines():
        val, name = (ln.split("|") + ["", ""])[:2]
        n = norm(val)
        if n and n not in contact_map:
            contact_map[n] = name
    return flat_map, contact_map

def main() -> int:
    ap = argparse.ArgumentParser(description="Cross-ref master directory; build Truecaller worklist.")
    ap.add_argument("--ingest", help="filled Truecaller worklist CSV to write names/emails back (dry-run unless --apply)")
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--real-ok", action="store_true")
    args = ap.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    if args.ingest:
        return ingest(Path(args.ingest), args.apply and args.real_ok)

    master = read_master()
    flat_map, contact_map = load_known()
    cross, worklist = [], []
    n_flat = n_contact = n_new = 0
    for m in master:
        ph = m["phone"]
        fm = flat_map.get(ph)
        cn = contact_map.get(ph)
        if fm:
            status = "known_with_flat"; n_flat += 1
        elif cn:
            status = "known_contact"; n_contact += 1
        else:
            status = "NEW_unknown"; n_new += 1
        cross.append({"serial": m["serial"], "phone": ph, "region": m["region"], "status": status,
                      "our_name": (fm or {}).get("name", "") or cn or "",
                      "our_flat": (fm or {}).get("flat", ""), "our_wing": (fm or {}).get("wing", ""),
                      "source_screenshot": m["source"]})
        if status == "NEW_unknown" and not ph.startswith("+"):  # Indian unknowns first (most actionable)
            worklist.append({"phone": ph, "region": m["region"], "truecaller_name": "", "email": "",
                             "flat_if_known": "", "notes": "", "source_screenshot": m["source"]})

    xref = OUT_DIR / "master_directory_crossref.csv"
    with xref.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(cross[0].keys())); w.writeheader(); w.writerows(cross)
    wl = OUT_DIR / "truecaller_worklist.csv"
    with wl.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["phone", "region", "truecaller_name", "email", "flat_if_known", "notes", "source_screenshot"])
        w.writeheader(); w.writerows(worklist)

    print(f"Master directory numbers: {len(master)}")
    print(f"  already known WITH a flat:  {n_flat}")
    print(f"  known as a contact (no flat): {n_contact}")
    print(f"  NEW / unknown to us:          {n_new}  (intl kept aside; Indian unknowns -> worklist: {len(worklist)})")
    print(f"\nCross-ref CSV:        {xref}")
    print(f"Truecaller worklist:  {wl}  ({len(worklist)} numbers to look up manually in the app)")
    return 0

def ingest(path: Path, write: bool) -> int:
    if not path.exists():
        print(f"Not found: {path}"); return 1
    rows = [r for r in csv.DictReader(path.open()) if r.get("phone") and (r.get("truecaller_name") or r.get("email"))]
    print(f"Filled rows to ingest: {len(rows)}")
    for r in rows[:12]:
        print(f"  {r['phone']}  name={r.get('truecaller_name','')!r}  email={r.get('email','')!r}  flat={r.get('flat_if_known','')!r}")
    if not write:
        print("\nDry run — pass --apply --real-ok to create/enrich contacts + contact_methods. (Ingest writer is a stub: "
              "wire to your contact upsert once the worklist is filled.)")
        return 0
    print("Apply path intentionally not auto-wired yet — confirm the contact-upsert target first.")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
