#!/usr/bin/env python3
"""
Scan all xlsx/xls/csv files on the drive for phone numbers and emails,
match to existing contacts, and insert missing contact_methods as 'unverified'.

Sources scanned:
  - /Volumes/RDH 5TB/RDH DATA 2024/RDH/RDH files/Excel Files/
  - /Volumes/RDH 5TB/RDH DATA 2024/RDH/RDH CONTACTS/

Match strategy:
  1. Phone exact match → contact via contact_methods.normalized_value or contacts.phone_primary
  2. Name fuzzy match (difflib ≥ 0.80) → contact.full_name

Only inserts if normalized_value NOT already present for that contact_id.
Dry-run by default. --apply to write.
"""
from __future__ import annotations
import argparse, csv, difflib, re, sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent

SCAN_DIRS = [
    Path("/Volumes/RDH 5TB/RDH DATA 2024/RDH/RDH files/Excel Files"),
    Path("/Volumes/RDH 5TB/RDH DATA 2024/RDH/RDH CONTACTS"),
]
PHASE = "file_scan_2026"
NAME_THRESH = 0.80

# ── db ────────────────────────────────────────────────────────────────────────
def get_conn():
    import psycopg2
    env = {}
    for line in (ROOT / "docker" / ".env").read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, _, v = line.partition("="); env[k.strip()] = v.strip()
    return psycopg2.connect(
        host=env.get("POSTGRES_HOST", "localhost"),
        port=int(env.get("POSTGRES_PORT", 5432)),
        dbname=env.get("POSTGRES_DB", "realdeal_os"),
        user=env.get("POSTGRES_USER", "realdeal_admin"),
        password=env.get("POSTGRES_PASSWORD", ""),
    )

# ── normalizers ───────────────────────────────────────────────────────────────
EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
PHONE_RE = re.compile(r"[\d\s\-\+\(\)\.]{7,20}")

def norm_phone(raw: str) -> str | None:
    if not raw: return None
    digits = re.sub(r"\D", "", str(raw))
    if digits.startswith("91") and len(digits) == 12: digits = digits[2:]
    if digits.startswith("0")  and len(digits) == 11: digits = digits[1:]
    return digits if len(digits) == 10 and digits[0] in "6789" else None

def norm_email(raw: str) -> str | None:
    if not raw: return None
    m = EMAIL_RE.search(str(raw))
    return m.group(0).lower() if m else None

def clean(v) -> str:
    return str(v).strip() if v is not None else ""

# ── file readers ──────────────────────────────────────────────────────────────
def rows_from_file(path: Path) -> list[dict]:
    """Return list of dicts with keys: name, phone, phone2, email, source_file, source_sheet, row_num"""
    suf = path.suffix.lower()
    results = []
    try:
        if suf in (".xlsx",):
            results = _read_xlsx(path)
        elif suf in (".xls",):
            results = _read_xls(path)
        elif suf in (".csv",):
            results = _read_csv(path)
    except Exception as e:
        print(f"  SKIP {path.name}: {e}")
    return results

def _sniff_cols(headers: list[str]) -> dict[str, int | None]:
    """Map semantic roles to column indices from header row."""
    h = [str(x).lower().strip() if x else "" for x in headers]
    def find(*keys) -> int | None:
        for k in keys:
            for i, hh in enumerate(h):
                if k in hh: return i
        return None
    return {
        "name":   find("name", "full name", "contact"),
        "phone":  find("phone", "mobile", "number", "contact no", "ph"),
        "phone2": find("phone 2", "phone2", "number 2", "alternate", "secondary", "mobile 2"),
        "email":  find("email", "mail", "e-mail"),
    }

def _make_row(cols: dict, values: list, source_file: str, source_sheet: str, row_num: int) -> dict | None:
    def get(key) -> str:
        idx = cols.get(key)
        return clean(values[idx]) if idx is not None and idx < len(values) else ""
    r = {"name": get("name"), "phone": get("phone"), "phone2": get("phone2"),
         "email": get("email"), "source_file": source_file, "source_sheet": source_sheet, "row_num": row_num}
    # skip entirely empty rows
    if not any([r["name"], r["phone"], r["email"]]): return None
    return r

def _read_xlsx(path: Path) -> list[dict]:
    import openpyxl
    results = []
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    for sh in wb.sheetnames:
        ws = wb[sh]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: continue
        # find header row (first row with ≥2 non-empty cells)
        hdr_idx = next((i for i, r in enumerate(rows) if sum(1 for c in r if c) >= 2), 0)
        cols = _sniff_cols(list(rows[hdr_idx]))
        if cols["name"] is None and cols["phone"] is None: continue
        for i, row in enumerate(rows[hdr_idx + 1:], hdr_idx + 2):
            r = _make_row(cols, list(row), path.name, sh, i)
            if r: results.append(r)
    wb.close()
    return results

def _read_xls(path: Path) -> list[dict]:
    import xlrd
    results = []
    wb = xlrd.open_workbook(str(path))
    for sh in wb.sheets():
        rows = [sh.row_values(i) for i in range(sh.nrows)]
        if not rows: continue
        hdr_idx = next((i for i, r in enumerate(rows) if sum(1 for c in r if clean(c)) >= 2), 0)
        cols = _sniff_cols(rows[hdr_idx])
        if cols["name"] is None and cols["phone"] is None: continue
        for i, row in enumerate(rows[hdr_idx + 1:], hdr_idx + 2):
            r = _make_row(cols, list(row), path.name, sh.name, i)
            if r: results.append(r)
    return results

def _read_csv(path: Path) -> list[dict]:
    results = []
    with open(path, encoding="utf-8-sig", errors="replace", newline="") as f:
        reader = csv.reader(line.replace("\x00", "") for line in f)
        rows = list(reader)
    if not rows: return results
    hdr_idx = next((i for i, r in enumerate(rows) if sum(1 for c in r if c.strip()) >= 2), 0)
    cols = _sniff_cols(rows[hdr_idx])
    if cols["name"] is None and cols["phone"] is None: return results
    for i, row in enumerate(rows[hdr_idx + 1:], hdr_idx + 2):
        r = _make_row(cols, row, path.name, "", i)
        if r: results.append(r)
    return results

# ── matching ──────────────────────────────────────────────────────────────────
def load_contacts(cur) -> tuple[dict, dict, dict]:
    """Returns (phone_to_contact, email_to_contact, id_to_name) dicts."""
    cur.execute("""
        SELECT c.id::text, c.full_name, c.phone_primary, c.email
        FROM contacts c WHERE c.canonical_status != 'merged'
    """)
    id_to_name, phone_map, email_map = {}, {}, {}
    for cid, name, ph, em in cur.fetchall():
        id_to_name[cid] = name or ""
        if ph:
            np = norm_phone(ph)
            if np: phone_map.setdefault(np, cid)
        if em:
            ne = norm_email(em)
            if ne: email_map.setdefault(ne, cid)

    cur.execute("""
        SELECT cm.contact_id::text, cm.normalized_value, cm.method_type
        FROM contact_methods cm
        WHERE cm.method_type IN ('mobile','phone','email') AND cm.normalized_value IS NOT NULL
          AND cm.validation_status != 'invalid'
    """)
    for cid, val, mtype in cur.fetchall():
        if cid not in id_to_name: continue
        if mtype in ("mobile", "phone"):
            phone_map.setdefault(val, cid)
        else:
            email_map.setdefault(val, cid)

    return phone_map, email_map, id_to_name

def load_existing_methods(cur) -> set[tuple[str, str]]:
    """Set of (contact_id, normalized_value) already in contact_methods."""
    cur.execute("""
        SELECT contact_id::text, normalized_value
        FROM contact_methods WHERE normalized_value IS NOT NULL AND contact_id IS NOT NULL
    """)
    return {(r[0], r[1]) for r in cur.fetchall()}

def fuzzy_match(name: str, id_to_name: dict) -> str | None:
    if not name or len(name) < 3: return None
    best, best_score = None, 0.0
    name_up = name.upper()
    for cid, cname in id_to_name.items():
        s = difflib.SequenceMatcher(None, name_up, cname.upper()).ratio()
        if s > best_score: best_score, best = s, cid
    return best if best_score >= NAME_THRESH else None

# ── main ──────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="Write to DB (default: dry run)")
    ap.add_argument("--no-name-match", action="store_true", help="Skip fuzzy name matching (fast)")
    args = ap.parse_args()

    conn = get_conn()
    cur = conn.cursor()

    print("Loading contacts from DB…")
    phone_map, email_map, id_to_name = load_contacts(cur)
    existing_methods = load_existing_methods(cur)
    print(f"  {len(id_to_name)} contacts | {len(phone_map)} phone keys | {len(email_map)} email keys")

    # ── scan files ────────────────────────────────────────────────────────────
    all_rows: list[dict] = []
    for scan_dir in SCAN_DIRS:
        for path in sorted(scan_dir.rglob("*")):
            if path.name.startswith("._"): continue
            if path.suffix.lower() not in (".xlsx", ".xls", ".csv"): continue
            rows = rows_from_file(path)
            if rows:
                print(f"  {len(rows):4d} rows  {path.name}")
                all_rows.extend(rows)

    print(f"\nTotal rows extracted: {len(all_rows)}")

    # ── match + dedupe candidates ─────────────────────────────────────────────
    to_insert: list[tuple] = []  # (contact_id, method_type, raw_value, norm_value, source_file, source_sheet, row_num)
    matched_by_phone = matched_by_name = unmatched = 0

    seen_inserts: set[tuple] = set()  # (contact_id, norm_value) — prevent same-run dupes

    for row in all_rows:
        phones = [norm_phone(row["phone"]), norm_phone(row["phone2"])]
        phones = [p for p in phones if p]
        email  = norm_email(row["email"])

        # find contact
        contact_id = None
        for p in phones:
            if p in phone_map:
                contact_id = phone_map[p]; matched_by_phone += 1; break
        if not contact_id and email and email in email_map:
            contact_id = email_map[email]; matched_by_phone += 1
        if not contact_id and not args.no_name_match:
            contact_id = fuzzy_match(row["name"], id_to_name)
            if contact_id: matched_by_name += 1
        if not contact_id:
            unmatched += 1; continue

        sf, ss, rn = row["source_file"], row["source_sheet"], row["row_num"]

        # queue new phones
        for p in phones:
            raw_ph = row["phone"] if norm_phone(row["phone"]) == p else row["phone2"]
            key = (contact_id, p)
            if key not in existing_methods and key not in seen_inserts:
                to_insert.append((contact_id, "mobile", raw_ph, p, sf, ss, rn))
                seen_inserts.add(key)

        # queue new email
        if email:
            key = (contact_id, email)
            if key not in existing_methods and key not in seen_inserts:
                to_insert.append((contact_id, "email", row["email"], email, sf, ss, rn))
                seen_inserts.add(key)

    print(f"\nMatch results:")
    print(f"  matched by phone/email : {matched_by_phone}")
    print(f"  matched by name        : {matched_by_name}")
    print(f"  unmatched              : {unmatched}")
    print(f"  new contact_methods    : {len(to_insert)}")

    if not to_insert:
        print("Nothing new to insert."); conn.close(); return

    # sample
    print(f"\nSample (first 10 to insert):")
    for cid, mtype, raw, norm, sf, ss, rn in to_insert[:10]:
        print(f"  {id_to_name.get(cid,'?')[:30]:30s}  {mtype:6s}  {norm}  ← {sf}")

    if not args.apply:
        print(f"\nDry run — {len(to_insert)} rows would be inserted. Add --apply to write.")
        conn.close(); return

    # ── insert ────────────────────────────────────────────────────────────────
    inserted = 0
    for cid, mtype, raw, norm, sf, ss, rn in to_insert:
        cur.execute("""
            INSERT INTO contact_methods
                (contact_id, method_type, raw_value, normalized_value,
                 validation_status, source_file, source_sheet, source_row_number,
                 raw_payload)
            VALUES (%s,%s,%s,%s,'unverified',%s,%s,%s,%s)
        """, [cid, mtype, raw, norm, sf, ss, rn,
              '{"source":"file_scan"}'])
        inserted += 1

    conn.commit()
    print(f"\nInserted {inserted} new contact_methods (validation_status=unverified).")
    conn.close()


if __name__ == "__main__":
    main()
