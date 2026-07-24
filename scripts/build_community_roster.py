#!/usr/bin/env python3
"""Work out exactly who must be saved to the sales phone and invited to the
broker WhatsApp Community — and never ask twice.

WhatsApp has no bulk-add: a person must be in the phone's address book, and
joining is by tapping an invite link. So this does the part that CAN be
automated — assemble the broker universe, diff it against the phone's own vCard
export, and emit two worklists:

  exports/community/brokers_to_save.vcf   → AirDrop to the phone, import once
  exports/community/brokers_in_phone.csv  → already saved; add/invite these now

State lives in community_roster (migration 072), so re-running is safe and
nobody gets re-invited.

  python3 scripts/build_community_roster.py --scan            (dry run)
  python3 scripts/build_community_roster.py --scan --apply
  python3 scripts/build_community_roster.py --mark-saved --apply
  python3 scripts/build_community_roster.py --status
"""
from __future__ import annotations

import argparse
import csv
import glob
import hashlib
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _db import run_psql, sql_literal as lit  # noqa: E402

PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTDIR = PROJECT_ROOT / "exports" / "community"
PHONEBOOK_GLOB = str(Path.home() / "Downloads" / "*.vcf")

# ponytail: truncated sha256, unsigned. It only tags a redirect and leaks
# nothing — same scheme as build_broker_channel_invites.py.
TOKEN_SECRET = "rdh-broker-community-v1"

TEL_RE = re.compile(r"^TEL[^:]*:(.+)$", re.I)
FN_RE = re.compile(r"^FN[^:]*:(.+)$", re.I)


def norm_phone(raw: object) -> str | None:
    """Indian mobile → +91XXXXXXXXXX, else None."""
    digits = re.sub(r"\D", "", str(raw or ""))
    if not digits:
        return None
    digits = digits.lstrip("0")
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    elif len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    if len(digits) != 10 or digits[0] not in "6789":
        return None
    return "+91" + digits


def q(sql: str) -> list[list[str]]:
    code, out = run_psql(sql)
    if code != 0:
        sys.exit(f"error: {out[:400]}")
    return [ln.split("|") for ln in out.splitlines() if ln]


def read_phonebook() -> dict[str, str]:
    """Every phone already in the sales phone, from its own vCard exports."""
    book: dict[str, str] = {}
    for path in glob.glob(PHONEBOOK_GLOB):
        try:
            text = Path(path).read_text(encoding="utf-8", errors="ignore")
        except OSError:
            continue
        name = ""
        for line in text.splitlines():
            line = line.strip()
            if line.upper().startswith("BEGIN:VCARD"):
                name = ""
            m = FN_RE.match(line)
            if m:
                name = m.group(1).strip()
                continue
            m = TEL_RE.match(line)
            if m:
                p = norm_phone(m.group(1))
                if p:
                    book.setdefault(p, name)
    return book


def broker_universe() -> dict[str, dict]:
    """Every broker we know, from the sources that actually identify brokers."""
    people: dict[str, dict] = {}

    def add(phone: str | None, name: str, source: str, detail: str) -> None:
        p = norm_phone(phone)
        if not p:
            return
        cur = people.get(p)
        # Prefer a real name over a bare number, and the earliest strong source.
        if cur is None:
            people[p] = {"name": name.strip(), "source": source, "detail": detail}
        elif not cur["name"] and name.strip():
            cur["name"] = name.strip()

    # 1. Members of the classified broker WhatsApp groups — the strongest signal:
    #    they are demonstrably active brokers already talking to us.
    for phone, name, chat in q("""
        SELECT m.phone, coalesce(m.display_name,''), coalesce(c.title,'')
          FROM wa_chat_members m
          JOIN wa_chats c ON c.beeper_chat_id = m.beeper_chat_id
         WHERE c.kind = 'broker_group' AND m.phone IS NOT NULL"""):
        add(phone, name, "wa_broker_group", chat)

    # 2. Anyone who has actually posted an offer in those groups.
    for phone, name in q("""
        SELECT sender_phone, coalesce(max(sender_name),'')
          FROM wa_market_offers WHERE sender_phone IS NOT NULL
         GROUP BY sender_phone"""):
        add(phone, name, "wa_offers", "posted a market offer")

    # 3. Broker spreadsheets on the drive, already catalogued.
    for (path,) in [(r[0],) for r in q("""
        SELECT file_path FROM drive_files
         WHERE NOT is_noise AND content_class = 'spreadsheet'
           AND doc_kind = 'broker_sheet'""")]:
        for name, phone in read_sheet_pairs(Path(path)):
            add(phone, name, "drive_sheet", Path(path).name)

    return people


def read_sheet_pairs(path: Path) -> list[tuple[str, str]]:
    """(name, phone) pairs out of a csv/xls/xlsx broker sheet. Never raises."""
    rows: list[list] = []
    try:
        ext = path.suffix.lower()
        if ext == ".csv":
            for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
                try:
                    with path.open(encoding=enc, newline="") as fh:
                        rows = list(csv.reader(fh))
                    break
                except UnicodeDecodeError:
                    continue
        elif ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            wb.close()
        elif ext == ".xls":
            import xlrd
            book = xlrd.open_workbook(str(path))
            sh = book.sheet_by_index(0)
            rows = [sh.row_values(i) for i in range(sh.nrows)]
    except Exception:  # noqa: BLE001 — a bad sheet must not stop the roster
        return []

    out: list[tuple[str, str]] = []
    for row in rows:
        phone = name = None
        for cell in row[:12]:
            p = norm_phone(cell)
            if p and not phone:
                phone = p
            elif not name and isinstance(cell, str) and len(cell.strip()) > 2 \
                    and not re.fullmatch(r"[\d\W]+", cell.strip()):
                name = cell.strip()[:80]
        if phone:
            out.append((name or "", phone))
    return out


def upsert(people: dict[str, dict], book: dict[str, str], apply: bool) -> dict:
    in_phone = {p for p in people if p in book}
    stats = {"total": len(people), "in_phonebook": len(in_phone),
             "to_save": len(people) - len(in_phone)}
    if not apply:
        return stats

    values = []
    for phone, meta in people.items():
        present = phone in book
        name = meta["name"] or book.get(phone, "") or ""
        values.append("(" + ", ".join([
            lit(phone), lit(name[:120]), lit("broker"), lit(meta["source"]),
            lit(meta["detail"][:200]), "TRUE" if present else "FALSE",
            lit("saved" if present else "to_save"),
            lit(hashlib.sha256((TOKEN_SECRET + phone).encode()).hexdigest()[:10]),
        ]) + ")")

    for i in range(0, len(values), 500):
        q(f"""
        INSERT INTO community_roster
          (phone, display_name, role, source, source_detail, in_phonebook, status, invite_token)
        VALUES {", ".join(values[i:i+500])}
        ON CONFLICT (phone) DO UPDATE SET
          display_name = coalesce(nullif(EXCLUDED.display_name,''), community_roster.display_name),
          in_phonebook = EXCLUDED.in_phonebook,
          -- never walk a person backwards: an invited/joined broker stays that way
          status = CASE WHEN community_roster.status IN ('invited','joined','declined','skip')
                        THEN community_roster.status ELSE EXCLUDED.status END,
          updated_at = now()""")
    return stats


def vcard(name: str, phone: str) -> str:
    label = f"RDH Broker · {name}" if name else f"RDH Broker · {phone}"
    return ("BEGIN:VCARD\r\nVERSION:3.0\r\n"
            f"FN:{label}\r\nN:{label};;;;\r\n"
            f"TEL;TYPE=CELL:{phone}\r\n"
            "CATEGORIES:RDH Broker\r\nEND:VCARD\r\n")


def emit() -> dict:
    OUTDIR.mkdir(parents=True, exist_ok=True)

    # Tier 1 first. The Community caps around 5,000 members and we know 8,000+
    # numbers, so "everyone" is not an option — and a broker who posts in our
    # groups is worth more than a name in a 2015 spreadsheet. Start narrow.
    active = q("""SELECT phone, coalesce(display_name,'') FROM community_roster
                   WHERE status = 'to_save' AND source IN ('wa_broker_group','wa_offers')
                   ORDER BY display_name NULLS LAST""")
    active_vcf = OUTDIR / "brokers_to_save_ACTIVE_FIRST.vcf"
    active_vcf.write_text("".join(vcard(n, p) for p, n in active), encoding="utf-8")

    to_save = q("""SELECT phone, coalesce(display_name,'') FROM community_roster
                    WHERE status = 'to_save' ORDER BY display_name NULLS LAST""")
    vcf = OUTDIR / "brokers_to_save.vcf"
    vcf.write_text("".join(vcard(n, p) for p, n in to_save), encoding="utf-8")

    # Named-only slice: importing 2,500 nameless numbers makes the phonebook
    # unusable and the saved name is what future parsing keys off.
    named = q("""SELECT phone, display_name FROM community_roster
                  WHERE status = 'to_save' AND coalesce(display_name,'') <> ''
                  ORDER BY display_name""")
    named_vcf = OUTDIR / "brokers_to_save_named_only.vcf"
    named_vcf.write_text("".join(vcard(n, p) for p, n in named), encoding="utf-8")

    ready = q("""SELECT phone, coalesce(display_name,''), coalesce(source,''), invite_token
                   FROM community_roster
                  WHERE in_phonebook AND status IN ('saved','to_save')
                  ORDER BY source, display_name""")
    csv_path = OUTDIR / "brokers_in_phone.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["phone", "name", "source", "invite_token"])
        w.writerows(ready)

    return {"active_first_vcf": f"{active_vcf.name} ({len(active)})",
            "named_only_vcf": f"{named_vcf.name} ({len(named)})",
            "all_to_save_vcf": f"{vcf.name} ({len(to_save)})",
            "ready_csv": f"{csv_path.name} ({len(ready)})",
            "outdir": str(OUTDIR)}


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--scan", action="store_true", help="rebuild the roster from all sources")
    ap.add_argument("--mark-saved", action="store_true",
                    help="after importing the vcf, flip to_save → saved")
    ap.add_argument("--status", action="store_true")
    ap.add_argument("--apply", action="store_true")
    a = ap.parse_args()

    if a.status:
        for row in q("SELECT status, people, in_phone FROM vw_community_roster_progress"):
            print(f"  {row[0]:<10} {row[1]:>6} people ({row[2]} already in phone)")
        return

    if a.mark_saved:
        if not a.apply:
            n = q("SELECT count(*) FROM community_roster WHERE status='to_save'")[0][0]
            print(f"DRY RUN — would mark {n} rows saved. Pass --apply.")
            return
        q("UPDATE community_roster SET status='saved', saved_at=now(), in_phonebook=TRUE, "
          "updated_at=now() WHERE status='to_save'")
        print("marked saved")
        return

    if a.scan:
        book = read_phonebook()
        print(f"phonebook: {len(book)} distinct mobile numbers")
        people = broker_universe()
        print(f"broker universe: {len(people)} distinct mobiles")
        stats = upsert(people, book, a.apply)
        print(f"in phonebook already: {stats['in_phonebook']}  |  need saving: {stats['to_save']}")
        if not a.apply:
            print("DRY RUN — pass --apply to write the roster and emit worklists.")
            return
        print(emit())
        return

    ap.print_help()


if __name__ == "__main__":
    main()
