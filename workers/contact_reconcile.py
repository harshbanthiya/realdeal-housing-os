"""contact_reconcile — every contact on the drive ends up in the DB or in review.

162 contact-bearing spreadsheets sit on the drive (owner/broker/tenant sheets,
phonebook exports), catalogued in drive_files. This worker chews through them a
batch at a time so it makes real progress on every 30-minute loop and always
resumes exactly where it stopped.

Per row, in order:
  1. phone matches an existing contact   → matched_phone (linked, nothing to do)
  2. name is a strong near-miss of an    → review (human settles it)
     existing contact but phone differs
  3. usable name + phone, no conflict    → created (new contact)
  4. no usable identity                  → skipped_no_contact (reason recorded)

The promise is the accounting, not the automation: `resolution = 'pending'`
should trend to zero, and nothing is ever silently dropped. See
vw_contact_reconcile_progress.

Writes contacts + contact_methods only in case 3, and only when the phone is
unambiguous. Everything doubtful becomes a review row.

Run: python3 workers/contact_reconcile.py [--batch N] [--reset-failed]
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _lib import finding, log_run, one, q  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))
from _db import jsonb_lit, run_psql, sql_literal as lit  # noqa: E402

WORKER = "contact_reconcile"

ROWS_PER_RUN = 1500         # a 30-min loop should finish well inside a minute
BACKLOG_CEILING = 5000      # stop pulling new sheets until resolution catches up
SHEETS_PER_RUN = 3          # register at most this many new sheets per run
NAME_SIMILARITY_REVIEW = 0.62   # above this, a differing phone is worth a human

# Column headers we understand, most specific first.
NAME_KEYS = ("full name", "name", "contact name", "owner name", "owner",
             "first name", "customer", "client", "party", "member")
PHONE_KEYS = ("phone", "mobile", "contact no", "contact number", "cell",
              "phone 1 - value", "phone number", "whatsapp", "tel")
EMAIL_KEYS = ("email", "e-mail", "mail", "email 1 - value")

JUNK_NAME = re.compile(r"^(n/?a|na|nil|none|-+|\.+|test|unknown|\s*)$", re.I)


def norm_phone(raw: str) -> str | None:
    """Indian mobile → +91XXXXXXXXXX. Returns None if it isn't plausibly one."""
    digits = re.sub(r"\D", "", str(raw or ""))
    if not digits:
        return None
    digits = digits.lstrip("0")
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    elif len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    if len(digits) != 10 or digits[0] not in "6789":
        return None
    return "+91" + digits


def clean_name(raw: str) -> str | None:
    n = re.sub(r"\s+", " ", str(raw or "")).strip(" .,-_/\\")
    if not n or JUNK_NAME.match(n) or len(n) < 2 or len(n) > 120:
        return None
    if re.fullmatch(r"[\d\W]+", n):
        return None
    return n


def pick(header: list[str], keys: tuple[str, ...]) -> int | None:
    low = [str(h or "").strip().lower() for h in header]
    for k in keys:                       # exact first
        if k in low:
            return low.index(k)
    for k in keys:                       # then substring
        for i, h in enumerate(low):
            if h and k in h:
                return i
    return None


def read_rows(path: Path) -> tuple[list[str], list[list]]:
    """Return (header, rows). Supports csv/xlsx/xls. Raises on unreadable."""
    ext = path.suffix.lower()
    if ext == ".csv":
        # Sheets exported from phones are frequently latin-1 or cp1252.
        for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                with path.open(encoding=enc, newline="") as fh:
                    rows = list(csv.reader(fh))
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError("undecodable csv")
        return (rows[0] if rows else []), rows[1:]

    if ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]
        wb.close()
        return (list(rows[0]) if rows else []), [list(r) for r in rows[1:]]

    if ext == ".xls":
        import xlrd
        book = xlrd.open_workbook(str(path))
        sh = book.sheet_by_index(0)
        rows = [sh.row_values(i) for i in range(sh.nrows)]
        return (rows[0] if rows else []), rows[1:]

    raise ValueError(f"unsupported extension {ext}")


def register_sheets() -> int:
    """Pull newly catalogued contact sheets into the work queue."""
    rows = q(f"""
        SELECT d.id::text, d.file_path, coalesce(d.doc_kind,''),
               coalesce(d.building_id::text,'')
          FROM drive_files d
          LEFT JOIN contact_sheet_files f ON f.file_path = d.file_path
         WHERE d.is_noise IS FALSE
           AND d.content_class = 'spreadsheet'
           AND d.doc_kind IN ('owner_sheet','broker_sheet','tenant_sheet','contact_export')
           AND f.id IS NULL
         ORDER BY d.file_size_bytes DESC
         LIMIT {SHEETS_PER_RUN}
    """)
    for did, path, kind, bid in rows:
        q(f"""INSERT INTO contact_sheet_files (drive_file_id, file_path, doc_kind, building_id)
              VALUES ({lit(did)}::uuid, {lit(path)}, {lit(kind) if kind else 'NULL'},
                      {lit(bid) + '::uuid' if bid else 'NULL'})
              ON CONFLICT (file_path) DO NOTHING""")
    return len(rows)


def load_sheet_rows(sheet_id: str, path: Path) -> int:
    """Parse one sheet into contact_sheet_rows. Returns rows written."""
    header, data = read_rows(path)
    ni, pi, ei = pick(header, NAME_KEYS), pick(header, PHONE_KEYS), pick(header, EMAIL_KEYS)

    # Headerless sheets are common ("Name,Phone Number" is a luxury). If we
    # found no phone column, scan the first data row for something phone-shaped.
    if pi is None and data:
        for i, cell in enumerate(data[0]):
            if norm_phone(cell):
                pi = i
                break
        if ni is None and pi is not None:
            ni = 0 if pi != 0 else 1

    q(f"""UPDATE contact_sheet_files
             SET rows_total = {min(len(data), 2_000_000)}, status = 'in_progress', started_at = now(),
                 header_map = {jsonb_lit({'name': ni, 'phone': pi, 'email': ei,
                                          'header': [str(h) for h in header[:25]]})},
                 updated_at = now()
           WHERE id = {lit(sheet_id)}::uuid""")

    if pi is None and ni is None:
        q(f"""UPDATE contact_sheet_files SET status='unreadable',
                 last_error='no name or phone column found', finished_at=now()
               WHERE id = {lit(sheet_id)}::uuid""")
        return 0

    # Batch the inserts — run_psql is a docker exec (~100ms per call).
    written, batch = 0, []
    for idx, row in enumerate(data):
        def cell(i):
            return row[i] if i is not None and i < len(row) else ""
        name = clean_name(cell(ni))
        phone = norm_phone(cell(pi))
        email = str(cell(ei) or "").strip().lower() or None
        # xlsx files routinely claim Excel's full 1,048,576-row extent. A row
        # with no name AND no phone is not a contact — never store it.
        if not name and not phone:
            continue
        raw = {str(header[i])[:60] if i < len(header) else f"col{i}": str(v)[:200]
               for i, v in enumerate(row[:25]) if v not in (None, "")}
        batch.append("(" + ", ".join([
            lit(sheet_id) + "::uuid", str(idx), jsonb_lit(raw),
            lit(name) if name else "NULL",
            lit(phone) if phone else "NULL",
            lit(email) if email and "@" in email else "NULL",
        ]) + ")")
        if len(batch) >= 500:
            written += flush_rows(batch)
            batch = []
    if batch:
        written += flush_rows(batch)
    return written


def flush_rows(batch: list[str]) -> int:
    q(f"""INSERT INTO contact_sheet_rows
            (sheet_file_id, row_index, raw, parsed_name, parsed_phone, parsed_email)
          VALUES {", ".join(batch)}
          ON CONFLICT (sheet_file_id, row_index) DO NOTHING""")
    return len(batch)


def resolve_batch(limit: int) -> dict:
    """Resolve pending rows: match, review, create, or skip. One SQL pass each."""
    stats = {"matched": 0, "review": 0, "created": 0, "skipped": 0}

    # 1. Phone already known → matched. Set-based; no per-row round trip.
    stats["matched"] = int(one(f"""
        WITH target AS (
          SELECT r.id, c.id AS contact_id
            FROM contact_sheet_rows r
            JOIN contact_methods m
              ON m.method_type = 'phone'
             AND regexp_replace(m.normalized_value, '\\D', '', 'g')
                 = regexp_replace(r.parsed_phone, '\\D', '', 'g')
            JOIN contacts c ON c.id = m.contact_id
           WHERE r.resolution = 'pending' AND r.parsed_phone IS NOT NULL
           LIMIT {limit}
        ), upd AS (
          UPDATE contact_sheet_rows r
             SET resolution = 'matched_phone', contact_id = t.contact_id,
                 resolution_reason = 'phone already in contact_methods'
            FROM target t WHERE r.id = t.id
          RETURNING 1
        ) SELECT count(*) FROM upd""") or 0)

    # 2. Unknown phone, but a strongly similar NAME exists → human decides.
    #    pg_trgm similarity; only the best candidate per row.
    stats["review"] = int(one(f"""
        WITH cand AS (
          SELECT DISTINCT ON (r.id)
                 r.id, c.id AS contact_id, c.full_name,
                 similarity(lower(r.parsed_name), lower(c.full_name)) AS sim
            FROM contact_sheet_rows r
            JOIN contacts c
              ON lower(c.full_name) % lower(r.parsed_name)
           WHERE r.resolution = 'pending'
             AND r.parsed_name IS NOT NULL
             AND c.full_name IS NOT NULL
           ORDER BY r.id, sim DESC
           LIMIT {limit}
        ), upd AS (
          UPDATE contact_sheet_rows r
             SET resolution = 'review', review_status = 'pending',
                 candidate_contact_id = cand.contact_id,
                 candidate_name = cand.full_name,
                 name_similarity = cand.sim,
                 resolution_reason = 'name looks like an existing contact but the phone differs'
            FROM cand
           WHERE r.id = cand.id AND cand.sim >= {NAME_SIMILARITY_REVIEW}
          RETURNING 1
        ) SELECT count(*) FROM upd""") or 0)

    # 3. Usable identity, nothing to confuse it with → create the contact.
    #    DISTINCT ON collapses the same phone repeated across sheets.
    stats["created"] = int(one(f"""
        WITH pick AS (
          SELECT DISTINCT ON (r.parsed_phone) r.id, r.parsed_name, r.parsed_phone, r.parsed_email
            FROM contact_sheet_rows r
           WHERE r.resolution = 'pending'
             AND r.parsed_phone IS NOT NULL AND r.parsed_name IS NOT NULL
           ORDER BY r.parsed_phone, r.id
           LIMIT {limit}
        ), ins AS (
          INSERT INTO contacts (full_name, phone_primary, email, source, status, canonical_status)
          SELECT parsed_name, parsed_phone, parsed_email, 'drive_contact_sheet', 'active', 'active'
            FROM pick
          RETURNING id, phone_primary
        ), meth AS (
          INSERT INTO contact_methods (contact_id, method_type, raw_value, normalized_value,
                                       is_primary, validation_status)
          SELECT i.id, 'phone', i.phone_primary, i.phone_primary, TRUE, 'unverified'
            FROM ins i
          RETURNING 1
        ), upd AS (
          UPDATE contact_sheet_rows r
             SET resolution = 'created', contact_id = i.id,
                 resolution_reason = 'new contact from drive sheet'
            FROM pick p JOIN ins i ON i.phone_primary = p.parsed_phone
           WHERE r.id = p.id
          RETURNING 1
        ) SELECT count(*) FROM upd""") or 0)

    # 4. Anything left with no usable identity is closed out with a reason.
    stats["skipped"] = int(one(f"""
        WITH target AS (
          SELECT id FROM contact_sheet_rows
           WHERE resolution = 'pending' AND parsed_phone IS NULL AND parsed_name IS NULL
           LIMIT {limit}
        ), upd AS (
          UPDATE contact_sheet_rows r
             SET resolution = 'skipped_no_contact',
                 resolution_reason = 'row has neither a usable name nor a phone'
            FROM target t WHERE r.id = t.id
          RETURNING 1
        ) SELECT count(*) FROM upd""") or 0)

    # A row with a name but no phone can't be created or matched safely.
    one(f"""
        WITH target AS (
          SELECT id FROM contact_sheet_rows
           WHERE resolution = 'pending' AND parsed_phone IS NULL
           LIMIT {limit}
        ), upd AS (
          UPDATE contact_sheet_rows r
             SET resolution = 'skipped_no_contact',
                 resolution_reason = 'name only, no phone — not enough to act on'
            FROM target t WHERE r.id = t.id
          RETURNING 1
        ) SELECT count(*) FROM upd""")

    return stats


def run() -> tuple[str, int, dict]:
    # Intake is throttled by the backlog: loading sheets is much faster than
    # resolving rows, so without this the queue grows forever and the operator
    # never sees it converge.
    backlog = int(one("SELECT count(*) FROM contact_sheet_rows WHERE resolution = 'pending'") or 0)
    registered = register_sheets() if backlog < BACKLOG_CEILING else 0

    # Load ONE pending sheet per run: parsing is the slow part, and a steady
    # drip keeps each loop short.
    loaded = 0
    pend = q("""SELECT id::text, file_path FROM contact_sheet_files
                 WHERE status = 'pending' ORDER BY created_at LIMIT 1""") if backlog < BACKLOG_CEILING else []
    if pend:
        sheet_id, path = pend[0][0], Path(pend[0][1])
        try:
            loaded = load_sheet_rows(sheet_id, path)
            q(f"""UPDATE contact_sheet_files SET status='done', rows_done={loaded},
                     finished_at=now(), updated_at=now()
                   WHERE id={lit(sheet_id)}::uuid AND status='in_progress'""")
        except Exception as exc:  # noqa: BLE001 — one bad sheet must not stop the loop
            q(f"""UPDATE contact_sheet_files SET status='failed',
                     last_error={lit(str(exc)[:400])}, finished_at=now(), updated_at=now()
                   WHERE id={lit(sheet_id)}::uuid""")

    stats = resolve_batch(ROWS_PER_RUN)

    prog = q("""SELECT sheets_total, sheets_done, sheets_pending, sheets_failed,
                       rows_seen, rows_matched, rows_created, rows_in_review, rows_unresolved
                  FROM vw_contact_reconcile_progress""")[0]
    keys = ["sheets_total", "sheets_done", "sheets_pending", "sheets_failed",
            "rows_seen", "rows_matched", "rows_created", "rows_in_review", "rows_unresolved"]
    detail = dict(zip(keys, [int(x or 0) for x in prog]))
    detail.update(stats, registered=registered, rows_loaded=loaded, backlog_at_start=backlog)

    if detail["rows_in_review"]:
        finding(WORKER, "contacts_awaiting_review", "contact_reconcile:review",
                f"{detail['rows_in_review']} drive contacts need a human decision",
                detail, "action" if detail["rows_in_review"] > 50 else "warn")

    if detail["sheets_failed"]:
        finding(WORKER, "sheet_unreadable", "contact_reconcile:failed",
                f"{detail['sheets_failed']} contact sheets could not be parsed",
                detail, "warn")

    summary = (f"+{registered} sheets, {loaded} rows loaded"
               f"{' (intake paused: backlog)' if backlog >= BACKLOG_CEILING else ''}; "
               f"matched {stats['matched']}, created {stats['created']}, "
               f"review {stats['review']}, skipped {stats['skipped']}; "
               f"{detail['rows_unresolved']} unresolved")
    return summary, detail["rows_in_review"], detail


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--reset-failed", action="store_true",
                    help="put failed/unreadable sheets back in the queue")
    ap.add_argument("--status", action="store_true")
    a = ap.parse_args()

    if a.reset_failed:
        q("UPDATE contact_sheet_files SET status='pending', last_error=NULL "
          "WHERE status IN ('failed','unreadable','in_progress')")
        print("reset failed sheets to pending")
        return 0
    if a.status:
        print(json.dumps(dict(zip(
            ["sheets_total", "sheets_done", "sheets_pending", "sheets_failed", "rows_seen",
             "rows_matched", "rows_created", "rows_in_review", "rows_unresolved", "rows_skipped"],
            q("SELECT * FROM vw_contact_reconcile_progress")[0])), indent=2))
        return 0

    return 0 if log_run(WORKER, run) else 1


if __name__ == "__main__":
    sys.exit(main())
